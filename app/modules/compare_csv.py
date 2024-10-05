from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
import csv
from openpyxl import load_workbook
from io import StringIO, BytesIO

def upload_files(request):
    return render(request, 'upload.html')

def normalize_column_names(headers):
    """
    Normalize column names by converting them to lowercase and stripping whitespace.
    """
    return [header.strip().lower() for header in headers]

def read_csv(file):
    """
    Read a CSV file and return the data as a list of dictionaries.
    """
    data = []
    file_content = StringIO(file.read().decode('utf-8'))
    reader = csv.DictReader(file_content)
    headers = normalize_column_names(reader.fieldnames)
    
    for row in reader:
        normalized_row = {key.strip().lower(): value for key, value in row.items()}
        data.append(normalized_row)
    
    return headers, data

def read_excel(file):
    """
    Read an Excel file and return the data as a list of dictionaries.
    """
    data = []
    workbook = load_workbook(filename=BytesIO(file.read()), data_only=True)
    sheet = workbook.active
    
    headers = normalize_column_names([cell.value for cell in sheet[1]])
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: value for i, value in enumerate(row)}
        data.append(row_data)
    
    return headers, data

def create_lookup_dict(data, columns):
    lookup_dict = {}
    for item in data:
        key = item[columns[0]]
        value = {col: item.get(col, '0') for col in columns[1:]}
        lookup_dict[key] = value
    return lookup_dict

# Function to clean keys
def clean_key(key, valid_keys):
    for valid_key in valid_keys:
        if valid_key in key:
            return valid_key
    return None

def compare_columns(request):
    try:
        if request.method == 'POST':
            # Get selected columns from POST request
            selected_columns_file1 = [col.lower() for col in request.POST.getlist('columns_file1')]
            selected_columns_file2 = [col.lower() for col in request.POST.getlist('columns_file2')]
            # Access uploaded files
            file1 = request.FILES.get('file1')
            file2 = request.FILES.get('file2')

            if not file1 or not file2:
                messages.error(request, 'Error: Missing files during comparison.')
                return redirect("upload_files")

            # Read CSV or Excel files
            headers1, data1 = None, None
            headers2, data2 = None, None

            if file1.name.endswith('.csv'):
                headers1, data1 = read_csv(file1)
            elif file1.name.endswith(('.xls', '.xlsx')):
                headers1, data1 = read_excel(file1)
            else:
                messages.error(request, 'Unsupported file format for file 1.')
                return redirect("upload_files")

            if file2.name.endswith('.csv'):
                headers2, data2 = read_csv(file2)
            elif file2.name.endswith(('.xls', '.xlsx')):
                headers2, data2 = read_excel(file2)
            else:
                messages.error(request, 'Unsupported file format for file 2.')
                return redirect("upload_files")

            updated_data1 = []
            updated_data2 = [] 
            for record in data1:
                updated_record = {}
                for key in record.keys():
                    clean_key_name = clean_key(key, selected_columns_file1)
                    if clean_key_name:
                        updated_record[clean_key_name] = record[key]
                updated_data1.append(updated_record)

            for record in data2:
                updated_record = {}
                for key in record.keys():
                    clean_key_name = clean_key(key, selected_columns_file1)
                    if clean_key_name:
                        updated_record[clean_key_name] = record[key]
                updated_data2.append(updated_record)

            file1_data = [{key: d[key] for key in selected_columns_file1 if key in d} for d in updated_data1]
            file2_data = [{key: d[key] for key in selected_columns_file2 if key in d} for d in updated_data2]

            file1_dict = create_lookup_dict(file1_data, selected_columns_file1)
            file2_dict = create_lookup_dict(file2_data, selected_columns_file1)

            # Find differing safety_ids or missing data
            differing_safety_ids = set()

            # Determine the key column
            key_column = selected_columns_file1[0]

            if len(selected_columns_file1) == 1:
                # Single column case: Check for presence in both files
                for key in file1_dict:
                    if key not in file2_dict:
                        differing_safety_ids.add(key)
                
                for key in file2_dict:
                    if key not in file1_dict:
                        differing_safety_ids.add(key)
                
                # Format response for single column case
                final_response = [{key_column: safety_id} for safety_id in differing_safety_ids]

            elif len(selected_columns_file1) == 2:
                # Two columns case: Compare values of the second column
                compare_column = selected_columns_file1[1]
                
                for safety_id in file1_dict:
                    if safety_id in file2_dict:
                        if file1_dict[safety_id].get(compare_column) != file2_dict[safety_id].get(compare_column):
                            differing_safety_ids.add(safety_id)
                
                # Check safety_ids that exist in file2 but not in file1
                for safety_id in file2_dict:
                    if safety_id not in file1_dict:
                        differing_safety_ids.add(safety_id)

                # Also include safety_ids present in file1 but missing in file2
                for safety_id in file1_dict:
                    if safety_id not in file2_dict:
                        differing_safety_ids.add(safety_id)

                # Format response for two columns case
                final_response = []
                for safety_id in differing_safety_ids:
                    response_entry = {col: file1_dict.get(safety_id, {}).get(col, '0') for col in selected_columns_file1[1:]}
                    response_entry[key_column] = safety_id
                    for col in selected_columns_file1[1:]:
                        response_entry[f'{col}_file2'] = file2_dict.get(safety_id, {}).get(col, '0')
                    final_response.append(response_entry)

            else:
                # More than two columns case: Compare all columns
                for safety_id in set(file1_dict) | set(file2_dict):
                    file1_values = file1_dict.get(safety_id, {})
                    file2_values = file2_dict.get(safety_id, {})
                    if file1_values != file2_values:
                        differing_safety_ids.add(safety_id)

                # Format response for more than two columns case
                final_response = []
                for safety_id in differing_safety_ids:
                    response_entry = {key_column: safety_id}
                    for col in selected_columns_file1[1:]:
                        response_entry[f'{col}_file1'] = file1_dict.get(safety_id, {}).get(col, '0')
                        response_entry[f'{col}_file2'] = file2_dict.get(safety_id, {}).get(col, '0')
                    final_response.append(response_entry)

            if final_response:
                # Return only the first key where there is a difference
                return render(request, 'differences.html', {
                    'differences': final_response,
                    'file1_name': file1.name,
                    'file2_name': file2.name
                })

            # If no differences are found
            messages.error(request, "No differences found.")
            return redirect("upload_files")
    except Exception as e:
        messages.error(request, e)
        return redirect("upload_files")
    messages.error(request, "Invalid request method.")
    return redirect("upload_files")
